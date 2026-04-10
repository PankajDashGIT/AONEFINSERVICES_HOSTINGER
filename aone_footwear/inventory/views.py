# inventory/views.py
from django.contrib.auth.models import User, Group
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
import csv
import io
import json
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP

from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (F, Subquery, OuterRef, DecimalField, Value, ExpressionWrapper, Sum, Min)
from django.db.models import Q
from django.db.models.functions import Coalesce
from django.db.models.functions import TruncMonth, TruncDate
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseForbidden
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.cache import cache_page
from xhtml2pdf import pisa

from .forms import LoginForm, PurchaseBillForm
from .forms import SupplierForm, SalesBillForm
from .models import (Category, Section, Size, PurchaseItem, Supplier, Color, Expense, PurchaseBill)
from .utils.whatsapp import send_whatsapp_document

def is_admin(user):
    return user.is_authenticated and user.groups.filter(name="ADMIN").exists()


def is_staff_user(user):
    print(user.is_authenticated, user.groups)
    return user.is_authenticated and user.groups.filter(name="STAFF").exists()


# Ensure WEASYPRINT_AVAILABLE is always defined to avoid unresolved reference warnings.
WEASYPRINT_AVAILABLE = False
try:
    # Importing WeasyPrint may raise ImportError or other exceptions if dependencies are missing.
    from weasyprint import HTML  # type: ignore

    WEASYPRINT_AVAILABLE = True
except Exception:
    WEASYPRINT_AVAILABLE = False

from .models import Product, SalesBill, SalesItem, Stock, Customer


# Helper to convert floats/strings to Decimal with 2 dp
def to_decimal(value):
    try:
        # Accept strings, floats or Decimals
        return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except Exception:
        return Decimal('0.00')


@login_required
def invoice_print(request, pk):
    """
    Invoice print/view endpoint.
    - If ?download=1 present, server generates a PDF (WeasyPrint) and returns it as attachment.
    - Else it renders the invoice_print.html template (same as current).
    """
    sale = get_object_or_404(SalesBill.objects.select_related('customer').prefetch_related('items__product'), pk=pk)

    if request.GET.get('download') in ['1', 'true', 'yes']:
        if not WEASYPRINT_AVAILABLE:
            return HttpResponse("Server-side PDF generation is not available. Install WeasyPrint.", status=500)

        # Render template to HTML string
        html_string = render_to_string('inventory/invoice_print.html', {'sale': sale})
        # Use absolute base URL so resources (images/css) can be loaded by WeasyPrint
        base_url = request.build_absolute_uri('/')

        # Generate PDF bytes
        pdf_file = HTML(string=html_string, base_url=base_url).write_pdf()

        # Respond with attachment
        filename = f"invoice-{sale.bill_number or sale.pk}.pdf"
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@cache_page(60 * 5)
def ledger_details(request, stock_id):
    # Get stock + product
    stock = get_object_or_404(Stock.objects.select_related("product"), id=stock_id)

    product = stock.product

    # -------------------------------
    # Latest PURCHASE (if any)
    # -------------------------------
    last_purchase = (PurchaseItem.objects.select_related("purchase").filter(product=product).order_by(
        "-purchase__bill_date").first())

    purchase_price = last_purchase.billing_price if last_purchase else 0
    purchase_date = (last_purchase.purchase.bill_date.strftime("%d-%m-%Y") if last_purchase else "-")

    # -------------------------------
    # Latest SALE (if any)
    # -------------------------------
    last_sale = (SalesItem.objects.select_related("sales_bill").filter(product=product).order_by(
        "-sales_bill__bill_date").first())

    sale_price = last_sale.selling_price if last_sale else 0
    sale_date = (last_sale.sales_bill.bill_date.strftime("%d-%m-%Y") if last_sale else "-")
    sale_invoice = (last_sale.sales_bill.bill_number if last_sale else "-")

    # -------------------------------
    # Profit
    # -------------------------------
    profit = sale_price - purchase_price

    return JsonResponse({"purchase_date": purchase_date, "purchase_price": round(purchase_price, 2),

        "sale_date": sale_date, "sale_price": round(sale_price, 2) if sale_price else "-", "sale_invoice": sale_invoice,
        # ✅ NEW

        "profit": round(profit, 2), })

def user_login(request):
    if request.user.is_authenticated:
        return redirect('retail_billing')

    form = LoginForm(request, data=request.POST or None)

    if request.method == 'POST' and form.is_valid():
        user = form.get_user()
        login(request, user)

        # 🔐 Ensure session key exists immediately
        request.session.save()

        return redirect('sales_dashboard')

    return render(request, 'inventory/login.html', {'form': form})


def user_logout(request):
    logout(request)
    return redirect('login')


# ---------------- ADD SUPPLIER ----------------
@login_required
# @user_passes_test(is_admin)
def supplier_add(request):
    form = SupplierForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("supplier_list")

    return render(request, "inventory/supplier_form.html", {"form": form, "title": "Add Supplier"})


# ---------------- EDIT SUPPLIER ----------------
@login_required
# @user_passes_test(is_admin)
def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    form = SupplierForm(request.POST or None, instance=supplier)

    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("supplier_list")

    return render(request, "inventory/supplier_form.html", {"form": form, "title": "Edit Supplier"})


# ---------------- LIST SUPPLIERS ----------------
@login_required
# @user_passes_test(is_admin)
def supplier_list(request):
    suppliers = Supplier.objects.all().order_by("name")

    supplier_data = []

    for supplier in suppliers:
        purchases = []

        # ✅ CORRECT RELATION NAME
        for p in supplier.purchasebill_set.all().order_by("-bill_date"):
            totals = p.items.aggregate(total_qty=Sum("quantity"), total_amt=Sum("line_total"), )

            purchases.append(
                {"bill_number": p.bill_number, "bill_date": p.bill_date, "total_qty": totals["total_qty"] or 0,
                    "total_amt": totals["total_amt"] or 0, })

        supplier_data.append({"supplier": supplier, "purchases": purchases, })

    return render(request, "inventory/supplier_list.html", {"supplier_data": supplier_data}, )


# ---------- Stock Purchase ----------
# updates only to purchase_view context to include categories/sections/sizes


# ... rest of file unchanged above purchase_view ...

# ---------- Stock Purchase ----------
@login_required
# @user_passes_test(is_admin)
@transaction.atomic
def purchase_view(request):
    form = PurchaseBillForm(request.POST or None)

    if request.method == "POST":

        # -----------------------------
        # Validate Bill Form
        # -----------------------------
        if not form.is_valid():
            return JsonResponse(
                {"error": "Invalid bill form", "details": form.errors},
                status=400
            )

        # -----------------------------
        # Read items JSON
        # -----------------------------
        items_json = request.POST.get("items_json")
        if not items_json:
            return JsonResponse({"error": "No items received"}, status=400)

        items = json.loads(items_json)
        if not items:
            return JsonResponse({"error": "No items in purchase"}, status=400)

        # -----------------------------
        # Save Bill Header
        # -----------------------------
        bill = form.save(commit=False)
        bill.created_by = request.user

        total_qty = sum(int(i["qty"]) for i in items)
        total_discount = sum(float(i.get("discount_rs") or 0) for i in items)

        total_gst = sum(
            ((float(i.get("price") or 0) * int(i.get("qty") or 0)) *
             float(i.get("gst_percent") or 0) / 100)
            for i in items
        )

        total_amount = sum(
            (float(i.get("price") or 0) * int(i.get("qty") or 0)) +
            ((float(i.get("price") or 0) * int(i.get("qty") or 0)) *
             float(i.get("gst_percent") or 0) / 100)
            for i in items
        )

        bill.total_qty = total_qty
        bill.total_discount = total_discount
        bill.total_gst = total_gst
        bill.total_amount = total_amount

        # Payment split
        if bill.payment_mode == "CASH":
            bill.cash_paid = total_amount
            bill.credit_amount = 0
        else:
            bill.cash_paid = 0
            bill.credit_amount = total_amount

        bill.save()

        # -----------------------------
        # Save Purchase Items
        # -----------------------------
        for item in items:

            # -------- REQUIRED FIELDS --------
            size_id = item.get("size_id")
            color_id = item.get("color_id") # or

            if not color_id:
                raise ValueError("Color is required for each purchase item")

            if not item.get("color_id"):
                raise ValueError(f"Missing color for item: {item}")

            # -------- Product --------
            product, _ = Product.objects.get_or_create(
                brand_id=item["brand_id"],
                category_id=item["category_id"],
                section_id=item["section_id"],
                size_id=size_id,
                color_id=color_id,
                defaults={
                    "mrp": item["mrp"],
                    "gst_percent": item.get("gst_percent", 0),
                }
            )

            # Always update pricing
            product.mrp = float(item.get("mrp") or 0)
            product.gst_percent = float(item.get("gst_percent") or 0)
            product.save()

            # -------- Calculations --------
            qty = int(item["qty"])
            billing_price = float(item["price"])
            disc_percent = float(item.get("discount_percent") or 0)
            disc_amount = float(item.get("discount_rs") or 0)
            gst_percent = float(item.get("gst_percent") or 0)
            mrp = float(item.get("mrp") or 0)
            msp = float(item.get("msp") or 0)

            gst_amount = ((billing_price * qty) * gst_percent) / 100
            line_total = (billing_price * qty) + gst_amount

            color_id = item.get("color_id")

            if not color_id:
                raise ValueError("Color is required for each purchase item")

            color = Color.objects.get(id=color_id)


            # -------- CREATE PURCHASE ITEM (COLOR IS MANDATORY) --------
            PurchaseItem.objects.create(
                purchase=bill,
                product=product,
                quantity=qty,
                mrp=mrp,
                billing_price=billing_price,
                discount_percent=disc_percent,
                discount_amount=disc_amount,
                gst_percent=gst_percent,
                gst_amount=gst_amount,
                line_total=line_total,
                msp=msp,
            )

            # -------- Stock Update --------
            stock, _ = Stock.objects.get_or_create(product=product)
            stock.quantity += qty
            stock.save()

        messages.success(request, "Purchase saved successfully!")
        return redirect("ledger")

    # -----------------------------
    # GET REQUEST
    # -----------------------------
    context = {
        "form": form,
        "brands": Brand.objects.all(),
        "suppliers": Supplier.objects.all(),
        "categories": Category.objects.all(),
        "sections": Section.objects.all(),
        "sizes": Size.objects.all(),
    }
    return render(request, "inventory/purchase.html", context)


@login_required
def post_login_redirect(request):
    if is_staff_user(request.user):
        return redirect("billing")
    return redirect("sales_dashboard")


@login_required
def invoice_view(request, bill_id):
    bill = get_object_or_404(SalesBill, pk=bill_id)
    return render(request, "inventory/invoice.html", {"bill": bill})


# ---------- Billing / POS ----------

@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
@transaction.atomic
def billing_view(request):
    """
    GET  : Render billing page
    POST : Validate → Create SalesBill & SalesItems → Update Stock
           → Redirect to Invoice HTML preview

    NOTES:
    - Billing is PURCHASE-ITEM (batch) based
    - GST is INCLUDED in selling price
    - Color is derived from Product (via PurchaseItem)
    """

    # =========================
    # GET : Render Billing Page
    # =========================
    if request.method == "GET":
        brands = Brand.objects.all().order_by("name")
        bill_form = SalesBillForm()
        return render(
            request,
            "inventory/billing.html",
            {
                "brands": brands,
                "bill_form": bill_form
            }
        )

    # =========================
    # POST : Submit Bill
    # =========================
    items_json = request.POST.get("items_json")
    if not items_json:
        return JsonResponse(
            {"success": False, "error": "No items supplied"},
            status=400
        )

    try:
        items = json.loads(items_json)
    except Exception:
        return JsonResponse(
            {"success": False, "error": "Invalid items_json"},
            status=400
        )

    if not isinstance(items, list) or not items:
        return JsonResponse(
            {"success": False, "error": "At least one item is required"},
            status=400
        )

    # =========================
    # CUSTOMER
    # =========================
    customer = None
    customer_mobile = request.POST.get("customer_mobile", "").strip()
    customer_name = request.POST.get("customer_name", "").strip()
    payment_mode = request.POST.get("payment_type", "CASH")

    if payment_mode == "CREDIT" and (not customer_mobile or not customer_name):
        return JsonResponse(
            {
                "success": False,
                "error": "Customer name and mobile are required for credit billing"
            },
            status=400
        )

    if customer_mobile:
        customer = Customer.objects.filter(phone=customer_mobile).first()

        if not customer:
            customer = Customer.objects.create(
                phone=customer_mobile,
                name=customer_name or "Customer"
            )

    if customer_name:
        customer = Customer.objects.filter(
            name=customer_name,
            phone=customer_mobile
        ).first()

        if not customer:
            customer = Customer.objects.create(
                name=customer_name,
                phone=customer_mobile or ""
            )

    if customer_name:
        customer = Customer.objects.filter(
            name=customer_name,
            phone=customer_mobile
        ).first()

        if not customer:
            customer = Customer.objects.create(
                name=customer_name,
                phone=customer_mobile or ""
            )

    # =========================
    # PRE-VALIDATION
    # =========================
    validated_items = []

    for idx, it in enumerate(items, start=1):
        try:
            purchase_item_id = int(it["purchase_item_id"])
            qty = int(it["qty"])
            price_unit = to_decimal(it["price"])
            gst_percent = to_decimal(it.get("gst_percent") or 0)
        except Exception:
            return JsonResponse(
                {"success": False, "error": f"Invalid data in item #{idx}"},
                status=400
            )

        if qty <= 0:
            return JsonResponse(
                {"success": False, "error": f"Invalid quantity in item #{idx}"},
                status=400
            )

        try:
            purchase_item = (
                PurchaseItem.objects
                .select_related("product", "product__color")
                .select_for_update()
                .get(pk=purchase_item_id)
            )
        except PurchaseItem.DoesNotExist:
            return JsonResponse(
                {"success": False, "error": f"Purchase batch not found (item #{idx})"},
                status=400
            )

        if purchase_item.quantity < qty:
            return JsonResponse(
                {
                    "success": False,
                    "error": (
                        f"Insufficient stock for "
                        f"{purchase_item.product} "
                        f"(Color {purchase_item.product.color}, "
                        f"MRP {purchase_item.mrp})"
                    )
                },
                status=400
            )

        mrp = to_decimal(purchase_item.mrp)

        # -------- MSP VALIDATION --------
        msp = to_decimal(getattr(purchase_item, "msp", 0))
        if msp > 0 and price_unit < msp:
            return JsonResponse(
                {
                    "success": False,
                    "error": (
                        f"Selling price ₹{price_unit} is below MSP ₹{msp} "
                        f"for {purchase_item.product}"
                    )
                },
                status=400
            )

        # -------- DISCOUNT LIMIT (15% of MRP) --------
        discount_per_unit = mrp - price_unit
        max_allowed = (mrp * Decimal("0.15")).quantize(Decimal("0.01"))

        if discount_per_unit > max_allowed:
            return JsonResponse(
                {
                    "success": False,
                    "error": (
                        f"Discount exceeds 15% of MRP for "
                        f"{purchase_item.product}"
                    )
                },
                status=400
            )

        validated_items.append({
            "purchase_item": purchase_item,
            "product": purchase_item.product,
            "qty": qty,
            "mrp": mrp,
            "price_unit": price_unit,
            "gst_percent": gst_percent,
        })

    # =========================
    # CREATE SALES BILL
    # =========================
    bill_number = timezone.now().strftime("BILL%Y%m%d%H%M%S")

    sales_bill = SalesBill.objects.create(
        bill_number=bill_number,
        bill_date=timezone.now(),
        customer=customer,
        payment_mode=payment_mode,
        total_qty=0,
        total_amount=Decimal("0.00"),
        total_payment=Decimal("0.00"),
        payment_received=Decimal("0.00"),
        balance_due=Decimal("0.00"),
        total_discount=Decimal("0.00"),
        total_gst=Decimal("0.00"),
        cgst=Decimal("0.00"),
        sgst=Decimal("0.00"),
        created_by=request.user
    )

    total_qty = 0
    total_amount = Decimal("0.00")
    total_gst = Decimal("0.00")
    total_discount = Decimal("0.00")

    # =========================
    # CREATE SALES ITEMS
    # =========================
    for item in validated_items:
        purchase_item = item["purchase_item"]
        product = item["product"]
        qty = item["qty"]
        mrp = item["mrp"]
        price_unit = item["price_unit"]
        gst_percent = item["gst_percent"]

        discount_per_unit = mrp - price_unit
        line_discount = (discount_per_unit * qty).quantize(Decimal("0.01"))

        line_total = (price_unit * qty).quantize(Decimal("0.01"))
        base_amount = (
            line_total * Decimal("100.00") /
            (Decimal("100.00") + gst_percent)
        ).quantize(Decimal("0.01"))
        line_gst = (line_total - base_amount).quantize(Decimal("0.01"))

        SalesItem.objects.create(
            sales_bill=sales_bill,
            product=product,             # Color flows via Product
            purchase_item=purchase_item,
            quantity=qty,
            mrp=mrp,
            selling_price=price_unit,
            discount_percent=(
                (discount_per_unit / mrp) * Decimal("100.00")
            ).quantize(Decimal("0.01")) if mrp > 0 else Decimal("0.00"),
            discount_amount=line_discount,
            gst_percent=gst_percent,
            gst_amount=line_gst,
            line_total=line_total
        )

        purchase_item.quantity -= qty
        purchase_item.save()

        total_qty += qty
        total_amount += line_total
        total_gst += line_gst
        total_discount += line_discount

    # =========================
    # FINALIZE BILL
    # =========================
    sales_bill.total_qty = total_qty
    sales_bill.total_amount = total_amount
    sales_bill.total_payment = total_amount
    sales_bill.payment_received = Decimal("0.00") if payment_mode == "CREDIT" else total_amount
    sales_bill.balance_due = total_amount - sales_bill.payment_received
    sales_bill.total_gst = total_gst
    sales_bill.total_discount = total_discount
    sales_bill.cgst = (total_gst / Decimal("2.00")).quantize(Decimal("0.01"))
    sales_bill.sgst = (total_gst / Decimal("2.00")).quantize(Decimal("0.01"))
    sales_bill.save()

    if customer and customer.phone:
        phone = customer.phone
        if not phone.startswith("91"):
            phone = "91" + phone

        invoice_url = request.build_absolute_uri(
            reverse("invoice_view", args=[sales_bill.id])
        )

        try:
            send_whatsapp_document(
                phone=phone,
                pdf_url=invoice_url,
                filename=f"Invoice-{sales_bill.bill_number}.pdf"
            )
        except Exception as e:
            print("WhatsApp failed:", e)

    # =========================
    # REDIRECT TO INVOICE
    # =========================
    invoice_url = reverse("invoice_view", args=[sales_bill.id])

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse(
            {"success": True, "invoice_url": invoice_url}
        )

    return HttpResponseRedirect(invoice_url)

@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
@transaction.atomic
def wholesale_billing_view(request):

    # ------------------------
    # GET
    # ------------------------
    if request.method == "GET":
        brands = Brand.objects.all().order_by("name")
        bill_form = SalesBillForm()

        return render(
            request,
            "inventory/wholesale_billing.html",
            {
                "brands": brands,
                "bill_form": bill_form
            }
        )

    # ------------------------
    # POST
    # ------------------------
    items_json = request.POST.get("items_json")

    if not items_json:
        return JsonResponse(
            {"success": False, "error": "No items supplied"},
            status=400
        )

    items = json.loads(items_json)

    customer = None
    customer_mobile = request.POST.get("customer_mobile", "").strip()
    customer_name = request.POST.get("customer_name", "").strip()
    payment_mode = request.POST.get("payment_type", "CASH")

    if payment_mode == "CREDIT" and (not customer_mobile or not customer_name):
        return JsonResponse(
            {
                "success": False,
                "error": "Customer name and mobile are required for credit billing"
            },
            status=400
        )

    if customer_mobile:
        customer = Customer.objects.filter(phone=customer_mobile).first()

        if not customer:
            customer = Customer.objects.create(
                phone=customer_mobile,
                name=customer_name or "Customer"
            )

    bill_number = timezone.now().strftime("WS%Y%m%d%H%M%S")

    sales_bill = SalesBill.objects.create(
        bill_number=bill_number,
        bill_date=timezone.now(),
        customer=customer,
        payment_mode=payment_mode,
        total_payment=Decimal("0.00"),
        payment_received=Decimal("0.00"),
        balance_due=Decimal("0.00"),
        created_by=request.user
    )

    total_qty = 0
    total_amount = Decimal("0")
    total_gst = Decimal("0")
    total_discount = Decimal("0")

    for it in items:

        purchase_item = PurchaseItem.objects.select_for_update().get(
            id=it["purchase_item_id"]
        )

        qty = int(it["qty"])
        price = Decimal(str(it["price"]))
        mrp = Decimal(str(it["mrp"]))
        gst = Decimal(str(it.get("gst_percent", 0)))

        if purchase_item.quantity < qty:
            return JsonResponse({
                "success": False,
                "error": "Insufficient stock"
            })

        line_total = price * qty

        base = line_total * Decimal("100") / (Decimal("100") + gst)
        gst_amount = line_total - base

        SalesItem.objects.create(
            sales_bill=sales_bill,
            product=purchase_item.product,
            purchase_item=purchase_item,
            quantity=qty,
            mrp=mrp,
            selling_price=price,
            gst_percent=gst,
            gst_amount=gst_amount,
            line_total=line_total
        )

        purchase_item.quantity -= qty
        purchase_item.save()

        total_qty += qty
        total_amount += line_total
        total_gst += gst_amount
        total_discount += (mrp - price) * qty

    sales_bill.total_qty = total_qty
    sales_bill.total_amount = total_amount
    sales_bill.total_payment = total_amount
    sales_bill.payment_received = Decimal("0.00") if payment_mode == "CREDIT" else total_amount
    sales_bill.balance_due = total_amount - sales_bill.payment_received
    sales_bill.total_gst = total_gst
    sales_bill.total_discount = total_discount
    sales_bill.cgst = total_gst / 2
    sales_bill.sgst = total_gst / 2
    sales_bill.save()

    return HttpResponseRedirect(
        reverse("invoice_view", args=[sales_bill.id])
    )


@login_required
def get_customer_by_mobile(request):
    mobile = request.GET.get("mobile", "").strip()

    if not mobile:
        return JsonResponse({"found": False})

    customer = Customer.objects.filter(phone=mobile).first()

    if not customer:
        return JsonResponse({"found": False})

    return JsonResponse({
        "found": True,
        "name": customer.name
    })



def api_product_info(request):
    size_id = request.GET.get('size_id')

    try:
        product = Product.objects.get(size_id=size_id)
        stock = Stock.objects.get(product=product)
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    except Stock.DoesNotExist:
        return JsonResponse({'error': 'Stock not found'}, status=404)

    data = {'mrp': product.mrp, 'stock_qty': stock.quantity, }
    return JsonResponse(data)


# ---------- Invoice PDF ----------
@login_required
def generate_invoice_pdf(request, bill_id):
    bill = get_object_or_404(SalesBill, pk=bill_id)

    html = render_to_string("inventory/invoice.html", {"bill": bill})

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (f'inline; filename="Invoice_{bill.bill_number}.pdf"')

    pisa.CreatePDF(html, dest=response)
    return response


# ---------- Stock Ledger ----------
@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
def ledger_view(request):
    """
    Stock Ledger (Batch-aware):
    - Filters
    - Derived stock (SUM of PurchaseItem.quantity)
    - Valuation from batch MRPs
    - Latest Supplier & Bill
    - Pagination
    """

    # --------------------------------------------------
    # BASE QUERY (PRODUCT-LEVEL)
    # --------------------------------------------------
    qs = (Product.objects.select_related("brand", "category", "section", "size").all())

    # --------------------------------------------------
    # FILTERS
    # --------------------------------------------------
    brand_id = request.GET.get("brand")
    category_id = request.GET.get("category")
    section_id = request.GET.get("section")
    size_id = request.GET.get("size")
    supplier_id = request.GET.get("supplier")

    if brand_id:
        qs = qs.filter(brand_id=brand_id)
    if category_id:
        qs = qs.filter(category_id=category_id)
    if section_id:
        qs = qs.filter(section_id=section_id)
    if size_id:
        qs = qs.filter(size_id=size_id)

    if supplier_id:
        qs = qs.filter(purchaseitem__purchase__supplier_id=supplier_id)

    qs = qs.distinct()

    # --------------------------------------------------
    # STOCK & VALUATION (FROM PURCHASEITEM)
    # --------------------------------------------------
    qs = qs.annotate(stock_qty=Coalesce(Sum("purchaseitem__quantity"), Value(0)), valuation=Coalesce(
        Sum(ExpressionWrapper(F("purchaseitem__quantity") * F("purchaseitem__mrp"),
            output_field=DecimalField(max_digits=14, decimal_places=2), )), Value(0),
        output_field=DecimalField(max_digits=14, decimal_places=2), ), )

    # --------------------------------------------------
    # LATEST PURCHASE DATA
    # --------------------------------------------------
    latest_pi = (PurchaseItem.objects.filter(product=OuterRef("pk")).order_by("-purchase__bill_date", "-purchase__id"))

    qs = qs.annotate(supplier_name=Subquery(latest_pi.values("purchase__supplier__name")[:1]),
        bill_number=Subquery(latest_pi.values("purchase__bill_number")[:1]),
        billing_price=Subquery(latest_pi.values("billing_price")[:1],
            output_field=DecimalField(max_digits=10, decimal_places=2), ), )

    # --------------------------------------------------
    # TOTALS
    # --------------------------------------------------
    totals = qs.aggregate(total_qty=Coalesce(Sum("stock_qty"), Value(0)),
        total_valuation=Coalesce(Sum("valuation"), Value(0),
            output_field=DecimalField(max_digits=18, decimal_places=2), ), )

    # --------------------------------------------------
    # PAGINATION
    # --------------------------------------------------
    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    # --------------------------------------------------
    # DROPDOWNS
    # --------------------------------------------------
    brands = Brand.objects.all()
    suppliers = Supplier.objects.all()

    categories = Category.objects.filter(brand_id=brand_id) if brand_id else Category.objects.none()
    sections = Section.objects.filter(category_id=category_id) if category_id else Section.objects.none()
    sizes = Size.objects.filter(section_id=section_id) if section_id else Size.objects.none()

    # --------------------------------------------------
    # CONTEXT
    # --------------------------------------------------
    context = {"stocks": page_obj, "page_obj": page_obj, "paginator": paginator, "brands": brands,
        "suppliers": suppliers, "categories": categories, "sections": sections, "sizes": sizes, "totals": totals, }

    return render(request, "inventory/ledger.html", context)


# ---------- AJAX APIs ----------

@login_required
def api_categories(request):
    brand_id = request.GET.get('brand_id')
    categories = Category.objects.filter(brand_id=brand_id).values('id', 'name')
    return JsonResponse(list(categories), safe=False)


@login_required
def api_sections(request):
    category_id = request.GET.get('category_id')
    sections = Section.objects.filter(category_id=category_id).values('id', 'name')
    return JsonResponse(list(sections), safe=False)


@login_required
def api_sizes(request):
    section_id = request.GET.get('section_id')
    sizes = Size.objects.filter(section_id=section_id).values('id', 'value')
    return JsonResponse(list(sizes), safe=False)

@login_required
def api_colors(request):
    size_id = request.GET.get("size_id")
    colors = Color.objects.filter(size_id=size_id).values("id", "value")
    return JsonResponse(list(colors), safe=False)

@login_required
def get_colors(request):
    """
    Returns available colors for billing screen
    Filtered by Brand + Category + Section + Size
    """

    brand_id = request.GET.get("brand")
    category_id = request.GET.get("category")
    section_id = request.GET.get("section")
    size_id = request.GET.get("size")

    if not all([brand_id, category_id, section_id, size_id]):
        return JsonResponse({"colors": []})

    colors = (
        Product.objects
        .filter(
            brand_id=brand_id,
            category_id=category_id,
            section_id=section_id,
            size_id=size_id,
            purchaseitem__quantity__gt=0
        )
        .values("color__id", "color__value")
        .distinct()
    )

    return JsonResponse({
        "colors": [
            {"id": c["color__id"], "name": c["color__value"]}
            for c in colors
        ]
    })


@login_required
def api_product_info(request):
    brand_id = request.GET.get('brand_id')
    category_id = request.GET.get('category_id')
    section_id = request.GET.get('section_id')
    size_id = request.GET.get('size_id')

    products = Product.objects.filter(brand_id=brand_id, category_id=category_id, section_id=section_id,
                                      size_id=size_id)

    # multi-MRP list + stock
    data = []
    for p in products:
        qty = getattr(p.stock, 'quantity', 0)
        data.append({'product_id': p.id, 'mrp': float(p.mrp), 'default_discount': float(p.default_discount_percent),
                     'gst_percent': float(p.gst_percent), 'stock_qty': qty, })

    return JsonResponse(data, safe=False)


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test

from .models import Brand
from .views import is_admin  # or import from where is_admin is defined


@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
def master_brand_add(request):
    """
    Add a new Brand.
    - Supports normal POST (redirect + messages)
    - Supports AJAX POST (JSON response)
    """

    if request.method != "POST":
        # Direct access not allowed
        return redirect("master_dashboard")

    name = request.POST.get("name", "").strip()

    # ---------- Validation ----------
    if not name:
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "message": "Brand name is required."}, status=400)
        messages.error(request, "Brand name is required.")
        return redirect("master_dashboard")

    # ---------- Duplicate Check ----------
    if Brand.objects.filter(name__iexact=name).exists():
        msg = f"Brand '{name}' already exists!"

        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "message": msg}, status=400)

        messages.warning(request, msg)
        return redirect("master_dashboard")

    # ---------- Save ----------
    Brand.objects.create(name=name)

    success_msg = f"Brand '{name}' added successfully!"

    # ---------- AJAX Response ----------
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True, "message": success_msg})

    # ---------- Normal Response ----------
    messages.success(request, success_msg)
    return redirect("master_dashboard")


@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
def master_category_add(request):
    if request.method == "POST":
        brand_id = request.POST.get("brand")
        name = request.POST.get("name", "").strip()

        if Category.objects.filter(brand_id=brand_id, name__iexact=name).exists():
            messages.warning(request, f"Category '{name}' already exists!")
        else:
            Category.objects.create(brand_id=brand_id, name=name)
            messages.success(request, f"Category '{name}' added successfully!")

    return redirect("master_dashboard")


@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
def master_section_add(request):
    if request.method == "POST":
        category_id = request.POST.get("category")
        name = request.POST.get("name", "").strip()

        if Section.objects.filter(category_id=category_id, name__iexact=name).exists():
            messages.warning(request, f"Section '{name}' already exists!")
        else:
            Section.objects.create(category_id=category_id, name=name)
            messages.success(request, f"Section '{name}' added successfully!")

    return redirect("master_dashboard")


@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
def master_size_add(request):
    if request.method == "POST":
        section = Section.objects.get(id=request.POST.get("section"))
        name = request.POST.get("name", "").strip()

        selected_sizes = request.POST.getlist("sizes")  # list: ["6","7","10","Free"]

        for size_val in selected_sizes:
            Size.objects.get_or_create(section=section, value=size_val)

        messages.success(request, "Sizes saved successfully!")
        return redirect("master_dashboard")

    return redirect("master_dashboard")

from django.db import IntegrityError

@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
def master_color_add(request):
    if request.method != "POST":
        return redirect("master_dashboard")

    size_id = request.POST.get("size")
    color_name = request.POST.get("name", "").strip()

    if not size_id:
        messages.error(request, "Size is required to add Color.")
        return redirect("master_dashboard")

    if not color_name:
        messages.error(request, "Color name is required.")
        return redirect("master_dashboard")

    try:
        Color.objects.create(
            size_id=size_id,
            value=color_name
        )
        messages.success(request, f"Color '{color_name}' added successfully!")
    except IntegrityError:
        messages.warning(
            request,
            f"Color '{color_name}' already exists for this size."
        )

    return redirect("master_dashboard")


@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
def master_dashboard(request):
    return render(request, "inventory/master/dashboard.html", {"brands": Brand.objects.all(), })





@login_required
@transaction.atomic
def sales_dashboard_view(request):
    """Render the Sales Dashboard shell — data loaded via AJAX."""
    return render(request, "inventory/sales_dashboard.html",
        {"is_admin": request.user.groups.filter(name="ADMIN").exists(),
            "is_staff": request.user.groups.filter(name="STAFF").exists(), })


@login_required
def sales_dashboard_data(request):

    start_str = request.GET.get("start_date")
    end_str = request.GET.get("end_date")
    search = request.GET.get("search", "").strip()

    page = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 10))

    today = timezone.localdate()
    month_start = today.replace(day=1)

    start_date = datetime.strptime(start_str, "%Y-%m-%d").date() if start_str else month_start
    end_date = datetime.strptime(end_str, "%Y-%m-%d").date() if end_str else today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    bills_qs = SalesBill.objects.select_related("customer").filter(
        bill_date__date__gte=start_date,
        bill_date__date__lte=end_date
    )

    if search:
        bills_qs = bills_qs.filter(
            Q(bill_number__icontains=search) |
            Q(customer__name__icontains=search) |
            Q(customer__phone__icontains=search)
        )

    # ---------------- KPIs ----------------

    today_sales = SalesBill.objects.filter(
        bill_date__date=today
    ).aggregate(total=Sum("total_amount"))["total"] or 0

    retail_sales = bills_qs.filter(
        bill_number__startswith="BILL"
    ).aggregate(total=Sum("total_amount"))["total"] or 0

    wholesale_sales = bills_qs.filter(
        bill_number__startswith="WS"
    ).aggregate(total=Sum("total_amount"))["total"] or 0

    totals = bills_qs.aggregate(
        total_sales=Sum("total_amount"),
        total_qty=Sum("total_qty")
    )

    total_sales = totals["total_sales"] or 0
    total_qty = totals["total_qty"] or 0

    last_7_days_start = today - timedelta(days=6)
    last_7_sales = SalesBill.objects.filter(
        bill_date__date__gte=last_7_days_start,
        bill_date__date__lte=today
    ).aggregate(total=Sum("total_amount"))["total"] or 0

    # ---------------- Payment Summary ----------------

    payments = [
        {"mode": p["payment_mode"], "amount": float(p["amount"] or 0)}
        for p in bills_qs.values("payment_mode").annotate(
            amount=Sum("total_amount")
        )
    ]

    trend_data = list(
        bills_qs.annotate(day=TruncDate("bill_date"))
        .values("day")
        .annotate(amount=Sum("total_amount"))
        .order_by("day")
    )

    pending_credit = bills_qs.filter(
        payment_mode="CREDIT"
    ).aggregate(total=Sum("balance_due"))["total"] or 0

    # ---------------- Pagination ----------------

    total_rows = bills_qs.count()
    start_idx = (page - 1) * page_size

    bills = bills_qs.order_by("-bill_date")[start_idx:start_idx + page_size]

    rows = []

    for bill in bills:

        customer_name = "Walk-in Customer"
        customer_mobile = "+91-xxxxxxxxxx"

        if bill.customer:
            customer_name = bill.customer.name or "Walk-in Customer"
            if bill.customer.phone:
                customer_mobile = f"+91-{bill.customer.phone}"

        rows.append({
            "bill_id": bill.id,
            "bill_no": bill.bill_number,
            "date": bill.bill_date.strftime("%d-%m-%Y"),
            "customer_name": customer_name,
            "customer_mobile": customer_mobile,
            "qty": bill.total_qty,
            "amount": float(bill.total_amount),
            "total_payment": float(bill.total_payment or bill.total_amount or 0),
            "payment_received": float(bill.payment_received or 0),
            "balance_due": float(bill.balance_due or 0),
            "payment": bill.get_payment_mode_display(),
            "can_collect_payment": bill.payment_mode == "CREDIT" and (bill.balance_due or 0) > 0,
        })

    return JsonResponse({

        "kpis": {
            "today_sales": float(today_sales),
            "last7_sales": float(last_7_sales),
            "retail_sales": float(retail_sales),
            "wholesale_sales": float(wholesale_sales),
            "total_sales": float(total_sales),
            "total_qty": int(total_qty),
            "pending_credit": float(pending_credit),
        },

        "payments": payments,
        "trend": [
            {
                "label": item["day"].strftime("%d-%m-%Y"),
                "amount": float(item["amount"] or 0)
            }
            for item in trend_data
        ],
        "table": {
            "rows": rows,
            "page": page,
            "page_size": page_size,
            "total_rows": total_rows,
            "total_pages": (total_rows + page_size - 1) // page_size,
        },

        "meta": {
            "start_date": start_date.strftime("%Y-%m-%d"),
            "end_date": end_date.strftime("%Y-%m-%d"),
        }

    })


@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
@transaction.atomic
def collect_bill_payment(request, bill_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid request method"}, status=405)

    bill = get_object_or_404(SalesBill.objects.select_for_update(), pk=bill_id)

    if bill.payment_mode != "CREDIT" or (bill.balance_due or Decimal("0.00")) <= 0:
        return JsonResponse({"success": False, "error": "This bill does not have any due payment"}, status=400)

    payment_mode = request.POST.get("payment_mode", "").strip().upper()
    amount_raw = request.POST.get("amount", "").strip()

    if payment_mode not in {"CASH", "UPI"}:
        return JsonResponse({"success": False, "error": "Payment mode must be Cash or UPI"}, status=400)

    try:
        amount = to_decimal(amount_raw)
    except Exception:
        amount = Decimal("0.00")

    if amount <= 0:
        return JsonResponse({"success": False, "error": "Enter a valid payment amount"}, status=400)

    current_due = to_decimal(bill.balance_due or 0)
    if amount > current_due:
        return JsonResponse({"success": False, "error": "Collected amount cannot exceed balance due"}, status=400)

    bill.payment_received = to_decimal((bill.payment_received or 0) + amount)
    bill.balance_due = to_decimal(current_due - amount)

    if bill.balance_due <= Decimal("0.00"):
        bill.balance_due = Decimal("0.00")
        bill.payment_mode = payment_mode

    bill.save(update_fields=["payment_received", "balance_due", "payment_mode", "updated_at"])

    customer_name = "Walk-in Customer"
    customer_mobile = "+91-xxxxxxxxxx"
    if bill.customer:
        customer_name = bill.customer.name or customer_name
        if bill.customer.phone:
            customer_mobile = f"+91-{bill.customer.phone}"

    return JsonResponse({
        "success": True,
        "message": "Payment collected successfully",
        "row": {
            "bill_id": bill.id,
            "bill_no": bill.bill_number,
            "date": bill.bill_date.strftime("%d-%m-%Y"),
            "customer_name": customer_name,
            "customer_mobile": customer_mobile,
            "qty": bill.total_qty,
            "amount": float(bill.total_amount),
            "total_payment": float(bill.total_payment or bill.total_amount or 0),
            "payment_received": float(bill.payment_received or 0),
            "balance_due": float(bill.balance_due or 0),
            "payment": bill.get_payment_mode_display(),
            "can_collect_payment": bill.payment_mode == "CREDIT" and (bill.balance_due or 0) > 0,
        }
    })


@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
def export_sales_excel(request):
    """
    Export filtered sales (same filters as dashboard) to CSV (Excel-friendly).
    URL: /sales/export/
    """

    start_str = request.GET.get("start_date")
    end_str = request.GET.get("end_date")
    search = request.GET.get("search", "").strip()

    today = timezone.localdate()
    month_start = today.replace(day=1)

    if start_str:
        start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
    else:
        start_date = month_start

    if end_str:
        end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
    else:
        end_date = today

    qs = (SalesItem.objects.select_related("sales_bill", "product__brand", "product__category", "product__section",
                                           "product__size").filter(sales_bill__bill_date__date__gte=start_date,
                                                                   sales_bill__bill_date__date__lte=end_date, ).order_by(
        "-sales_bill__bill_date", "-id"))

    if search:
        qs = qs.filter(Q(sales_bill__bill_number__icontains=search) | Q(product__brand__name__icontains=search) | Q(
            product__category__name__icontains=search) | Q(product__section__name__icontains=search) | Q(
            product__size__value__icontains=search))

    # CSV response
    response = HttpResponse(content_type="text/csv")
    filename = f"sales_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # ===========================
    #   UPDATED COLUMN HEADERS
    # ===========================
    writer.writerow(["Date", "Bill No", "Article", "Category", "Size", "Qty", "MRP", "Discount", "Total GST", "Total",
                     "Payment Mode", "Customer Mobile Number"])
    for item in qs:
        bill = item.sales_bill
        product = item.product

        line_amount = float(item.line_total or (item.quantity * item.selling_price))
        discount = float(item.discount_amount or 0)

        # Compute total GST
        total_gst = float(item.gst_amount)

        writer.writerow([bill.bill_date.strftime("%d-%m-%Y"), bill.bill_number,
                         product.article_no or f"{product.brand.name}/{product.section.name}/{product.size.value}",
                         product.category.name, product.size.value, item.quantity, f"{item.mrp:.2f}", f"{discount:.2f}",
                         f"{total_gst:.2f}",  # NEW COLUMN
                         f"{line_amount:.2f}", bill.payment_mode, bill.customer.phone if bill.customer else ""])

    return response


def landing_view(request):
    return render(request, "inventory/landing.html")


def privacy_view(request):
    return render(request, "inventory/privacy.html")


def terms_view(request):
    return render(request, "inventory/terms.html")


@login_required()
def contact_view(request):
    return render(request, "inventory/contact_us.html")


@login_required
def check_purchase_bill(request):
    supplier_id = request.GET.get("supplier_id")
    bill_number = request.GET.get("bill_number", "").strip()

    if not supplier_id or not bill_number:
        return JsonResponse({"exists": False})

    exists = PurchaseBill.objects.filter(supplier_id=supplier_id, bill_number__iexact=bill_number).exists()

    return JsonResponse({"exists": exists})


@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
def party_wise_purchase_view(request):
    suppliers = Supplier.objects.all().order_by("name")

    supplier_id = request.GET.get("supplier")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    items = PurchaseItem.objects.select_related("purchase", "product__brand", "product__category", "product__section",
                                                "product__size", "purchase__supplier")

    if supplier_id:
        items = items.filter(purchase__supplier_id=supplier_id)

    if start_date:
        items = items.filter(purchase__bill_date__date__gte=start_date)

    if end_date:
        items = items.filter(purchase__bill_date__date__lte=end_date)

    items = items.order_by("-purchase__bill_date")

    context = {"suppliers": suppliers, "items": items, "selected_supplier": supplier_id, "start_date": start_date,
               "end_date": end_date, }

    return render(request, "inventory/reports/party_wise_purchase.html", context)


@login_required
@user_passes_test(lambda u: is_admin(u) or is_staff_user(u))
def export_stock_ledger_csv(request):
    """
    Export Stock Ledger to CSV (respects current filters)
    """

    # --------------------------------------------------
    # BASE QUERY
    # --------------------------------------------------
    qs = (Stock.objects.select_related("product", "product__brand", "product__category", "product__section",
        "product__size", ).all())

    # --------------------------------------------------
    # APPLY FILTERS
    # --------------------------------------------------
    brand_id = request.GET.get("brand")
    category_id = request.GET.get("category")
    section_id = request.GET.get("section")
    size_id = request.GET.get("size")
    supplier_id = request.GET.get("supplier")

    if brand_id:
        qs = qs.filter(product__brand_id=brand_id)
    if category_id:
        qs = qs.filter(product__category_id=category_id)
    if section_id:
        qs = qs.filter(product__section_id=section_id)
    if size_id:
        qs = qs.filter(product__size_id=size_id)
    if supplier_id:
        qs = qs.filter(product__purchaseitem__purchase__supplier_id=supplier_id).distinct()

    # --------------------------------------------------
    # VALUATION (MRP × STOCK)
    # --------------------------------------------------
    valuation_expr = ExpressionWrapper(F("quantity") * F("product__mrp"),
        output_field=DecimalField(max_digits=14, decimal_places=2), )

    qs = qs.annotate(
        valuation=Coalesce(valuation_expr, Value(0), output_field=DecimalField(max_digits=14, decimal_places=2), ))

    # --------------------------------------------------
    # LATEST PURCHASE DATA
    # --------------------------------------------------
    latest_pi = (
        PurchaseItem.objects.filter(product=OuterRef("product")).order_by("-purchase__bill_date", "-purchase__id"))

    qs = qs.annotate(supplier_name=Subquery(latest_pi.values("purchase__supplier__name")[:1]),
        bill_number=Subquery(latest_pi.values("purchase__bill_number")[:1]),
        billing_price=Subquery(latest_pi.values("billing_price")[:1],
            output_field=DecimalField(max_digits=10, decimal_places=2), ), )

    # --------------------------------------------------
    # CREATE CSV RESPONSE
    # --------------------------------------------------
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="stock_ledger.csv"'

    writer = csv.writer(response)

    # HEADER
    writer.writerow(
        ["Brand", "Category", "Section", "Size", "Current Stock", "Billing Price", "MRP Valuation", "Supplier",
            "Bill No", ])

    # ROWS
    for s in qs:
        writer.writerow(
            [s.product.brand.name, s.product.category.name, s.product.section.name, s.product.size.value, s.quantity,
                float(s.billing_price) if s.billing_price else "", float(s.valuation) if s.valuation else "",
                s.supplier_name or "", s.bill_number or "", ])

    return response



def is_admin(user):
    return user.groups.filter(name="ADMIN").exists() or user.is_superuser


@login_required
@user_passes_test(is_admin)
def staff_management_view(request):
    staff_group = Group.objects.get(name="STAFF")

    users = (User.objects.filter(groups=staff_group).order_by("username"))

    return render(request, "admin/staff_management.html", {"users": users})


@login_required
@user_passes_test(is_admin)
def staff_add_edit_view(request, user_id=None):
    staff_group = Group.objects.get(name="STAFF")
    user = get_object_or_404(User, id=user_id) if user_id else None

    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST.get("email", "")
        password = request.POST.get("password")

        if user:
            user.username = username
            user.email = email
            if password:
                user.set_password(password)
            user.save()
        else:
            user = User.objects.create_user(username=username, email=email, password=password)
            user.groups.add(staff_group)

        messages.success(request, "Staff saved successfully")
        return redirect("staff_management")

    return render(request, "admin/staff_form.html", {"user": user})


@login_required
@user_passes_test(is_admin)
def staff_toggle_active(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = not user.is_active
    user.save()
    return redirect("staff_management")


@login_required
def expense_management_view(request):
    qs = Expense.objects.all()

    # Date filters
    start = request.GET.get("start_date")
    end = request.GET.get("end_date")

    if start and end:
        qs = qs.filter(expense_date__range=[start, end])

    # STAFF → only their expenses
    if not is_admin(request.user):
        qs = qs.filter(created_by=request.user)

    qs = qs.order_by("-expense_date")

    return render(request, "inventory/expense.html",
                  {"expenses": qs, "categories": Expense.CATEGORY_CHOICES, "is_admin": is_admin(request.user), })


@login_required
# @user_passes_test(is_admin)
def expense_chart_data(request):
    today = timezone.localdate()
    start_date = today - timedelta(days=6)

    qs = (
        Expense.objects.filter(expense_date__gte=start_date).values("category").annotate(total=Sum("amount")).order_by(
            "category"))

    return JsonResponse({"labels": [q["category"] for q in qs], "data": [float(q["total"]) for q in qs], })


@login_required
def expense_add(request):
    if request.method == "POST":
        Expense.objects.create(category=request.POST["category"], description=request.POST.get("description", ""),
            amount=request.POST["amount"], created_by=request.user, approved=request.user.is_superuser,
            # Admin auto-approved
        )
        return redirect("expense_management")


@login_required
@user_passes_test(is_admin)
def expense_approve(request, expense_id):
    exp = get_object_or_404(Expense, id=expense_id)
    exp.approved = True
    exp.save()
    return redirect("expense_management")


@login_required
# @user_passes_test(is_admin)
def monthly_expense_report(request):
    data = (Expense.objects.filter(approved=True).annotate(month=TruncMonth("expense_date")).values("month").annotate(
        total=Sum("amount")).order_by("-month"))
    return JsonResponse(list(data), safe=False)


@login_required
def expense_edit(request, expense_id):
    exp = get_object_or_404(Expense, id=expense_id)

    if exp.created_by != request.user and not is_admin(request.user):
        return HttpResponseForbidden()

    if request.method == "POST":
        exp.category = request.POST["category"]
        exp.description = request.POST["description"]
        exp.amount = request.POST["amount"]
        exp.approved = False  # re-approval required
        exp.save()
        return redirect("expense_management")


@login_required
# @user_passes_test(is_admin)
def export_expenses_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="expenses.csv"'

    writer = csv.writer(response)
    writer.writerow(["Date", "Category", "Description", "Amount", "User", "Approved"])

    for e in Expense.objects.all():
        writer.writerow([e.expense_date, e.category, e.description, e.amount, e.created_by.username,
            "Yes" if e.approved else "No", ])

    return response


@login_required
# @user_passes_test(is_admin)
def profit_dashboard_data(request):
    total_sales = (SalesBill.objects.aggregate(total=Sum("total_amount"))["total"] or 0)

    total_expenses = (Expense.objects.filter(approved=True).aggregate(total=Sum("amount"))["total"] or 0)

    return JsonResponse({"sales": float(total_sales), "expenses": float(total_expenses),
        "profit": float(total_sales - total_expenses), })


@login_required
# @user_passes_test(is_admin)
def expense_delete(request, expense_id):
    Expense.objects.filter(id=expense_id).delete()
    return redirect("expense_management")


@login_required
def get_product_mrps(request):
    product_id = request.GET.get("product_id")
    if not product_id:
        return JsonResponse({"results": []})

    qs = (
        PurchaseItem.objects.filter(product_id=product_id, quantity__gt=0).values("mrp", "gst_percent", "msp").annotate(
            stock=Sum("quantity"), purchase_item_id=Min("id")).order_by("mrp"))

    data = []
    for row in qs:
        data.append(
            {"purchase_item_id": row["purchase_item_id"], "mrp": float(row["mrp"]), "gst": float(row["gst_percent"]),
                "default_disc": 10.0, "msp": float(row["msp"] or 0), "stock": int(row["stock"]), })

    return JsonResponse({"results": data})


@login_required
def get_product_id(request):
    brand_id = request.GET.get("brand")
    category_id = request.GET.get("category")
    section_id = request.GET.get("section")
    size_id = request.GET.get("size")
    color_id = request.GET.get("color")

    if not all([brand_id, category_id, section_id, size_id, color_id]):
        return JsonResponse({"product_id": None})

    product = Product.objects.filter(
        brand_id=brand_id,
        category_id=category_id,
        section_id=section_id,
        size_id=size_id,
        color_id=color_id
    ).first()

    return JsonResponse({"product_id": product.id if product else None})


@login_required
def ledger_product_details(request, product_id):
    # Latest purchase (batch-aware)
    latest_pi = (
        PurchaseItem.objects.filter(product_id=product_id).select_related("purchase").order_by("-purchase__bill_date",
                                                                                               "-purchase__id").first())

    # Latest sale
    latest_si = (
        SalesItem.objects.filter(product_id=product_id).select_related("sales_bill").order_by("-sales_bill__bill_date",
                                                                                              "-sales_bill__id").first())

    return JsonResponse({"purchase_date": (latest_pi.purchase.bill_date.strftime("%d-%m-%Y") if latest_pi else None),
        "purchase_price": (
            float(latest_pi.billing_price) if latest_pi and latest_pi.billing_price is not None else None),
        "sale_date": (latest_si.sales_bill.bill_date.strftime("%d-%m-%Y") if latest_si else None),
        "sale_price": (float(latest_si.selling_price) if latest_si else None),
        "sale_invoice": (latest_si.sales_bill.bill_number if latest_si else None), "profit": (float(
            latest_si.selling_price - latest_pi.billing_price) if latest_si and latest_pi and latest_pi.billing_price is not None else None), })


@login_required
# @user_passes_test()
@transaction.atomic
def import_purchase_file(request):

    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid request"}, status=400)

    upload = request.FILES.get("file")
    if not upload:
        return JsonResponse({"success": False, "error": "No file uploaded"}, status=400)

    items = []
    errors = []

    try:
        # -------------------------------
        # READ FILE (CSV / XLSX)
        # -------------------------------
        if upload.name.lower().endswith(".csv"):
            rows = csv.DictReader(io.StringIO(upload.read().decode("utf-8")))
        else:
            import openpyxl
            wb = openpyxl.load_workbook(upload)
            ws = wb.active
            headers = [str(c.value).strip() for c in ws[1]]
            rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

        # -------------------------------
        # PROCESS ROWS
        # -------------------------------
        for row_no, row in enumerate(rows, start=2):
            try:
                brand_name = str(row.get("Brand", "")).strip()
                category_name = str(row.get("Category", "")).strip()
                section_name = str(row.get("Section", "")).strip()
                size_value = str(row.get("Size", "")).strip()
                color_value = str(row.get("Color", "")).strip()

                if not all([brand_name, category_name, section_name, size_value, color_value]):
                    raise ValueError("Brand, Category, Section, Size and Color are required")

                mrp = Decimal(row.get("MRP", 0))
                price = Decimal(row.get("BillingPrice", 0))
                qty = int(row.get("Qty", 0))
                gst = Decimal(row.get("GST", 0))

                if qty <= 0:
                    raise ValueError("Quantity must be > 0")

                # -------------------------------
                # AUTO CREATE MASTER DATA
                # -------------------------------
                brand, _ = Brand.objects.get_or_create(
                    name__iexact=brand_name,
                    defaults={"name": brand_name}
                )

                category, _ = Category.objects.get_or_create(
                    brand=brand,
                    name__iexact=category_name,
                    defaults={"name": category_name}
                )

                section, _ = Section.objects.get_or_create(
                    category=category,
                    name__iexact=section_name,
                    defaults={"name": section_name}
                )

                size, _ = Size.objects.get_or_create(
                    section=section,
                    value__iexact=size_value,
                    defaults={"value": size_value}
                )

                # 🔴 AUTO CREATE COLOR (THIS IS NEW)
                color, _ = Color.objects.get_or_create(
                    size=size,
                    value__iexact=color_value,
                    defaults={"value": color_value}
                )

                # -------------------------------
                # CALCULATIONS
                # -------------------------------
                discount_rs = mrp - price
                discount_percent = (discount_rs / mrp * 100) if mrp else Decimal("0")
                gst_amount = ((price * qty) * gst) / 100
                line_total = (price * qty) + gst_amount
                msp = price + (price * Decimal("0.20"))

                # -------------------------------
                # SEND BACK TO UI
                # -------------------------------
                items.append({
                    "brand_id": brand.id,
                    "category_id": category.id,
                    "section_id": section.id,
                    "size_id": size.id,

                    # 🔴 ALWAYS PRESENT
                    "color_id": color.id,
                    "color_name": color.value,

                    "brand_name": brand.name,
                    "category_name": category.name,
                    "section_name": section.name,
                    "size_name": size.value,

                    "mrp": float(mrp),
                    "price": float(price),
                    "qty": qty,
                    "gst_percent": float(gst),

                    "discount_rs": float(discount_rs),
                    "discount_percent": float(discount_percent),
                    "gst_amount": float(gst_amount),
                    "line_total": float(line_total),
                    "msp": float(msp),
                })

            except Exception as e:
                errors.append(f"Row {row_no}: {str(e)}")

        if errors:
            return JsonResponse({"success": False, "errors": errors})

        return JsonResponse({"success": True, "items": items})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@login_required
def profit_dashboard_api(request):

    today = timezone.localdate()

    range_type = request.GET.get("range", "today")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # ---------------- DATE RANGE ----------------
    if range_type == "7":
        start = today - timedelta(days=6)
        end = today
    elif range_type == "15":
        start = today - timedelta(days=14)
        end = today
    elif range_type == "custom" and start_date and end_date:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
    else:
        start = today
        end = today

    # Convert DATE → DATETIME (FULL DAY RANGE)
    start_dt = timezone.make_aware(datetime.combine(start, datetime.min.time()))
    end_dt = timezone.make_aware(datetime.combine(end, datetime.max.time()))

    # ---------------- SALES ----------------
    sales_qs = SalesItem.objects.filter(
        sales_bill__bill_date__range=(start_dt, end_dt)
    )

    total_sales = sales_qs.aggregate(
        total=Sum("line_total")
    )["total"] or 0

    # ---------------- PURCHASE COST ----------------
    purchase_cost_expr = ExpressionWrapper(
        F("quantity") * F("purchase_item__billing_price"),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )

    total_purchase_cost = sales_qs.aggregate(
        total=Sum(purchase_cost_expr)
    )["total"] or 0

    # ---------------- EXPENSES ----------------
    total_expenses = Expense.objects.filter(
        approved=True,
        expense_date__range=(start_dt, end_dt)
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0

    # ---------------- PROFIT ----------------
    profit = total_sales - total_purchase_cost - total_expenses

    return JsonResponse({
        "range": {
            "start": start.strftime("%d-%m-%Y"),
            "end": end.strftime("%d-%m-%Y"),
        },
        "total_sales": float(total_sales),
        "purchase_cost": float(total_purchase_cost),
        "expenses": float(total_expenses),
        "profit": float(profit),
    })

@login_required
def bill_wise_profit_api(request):
    """
    Bill-wise profit for a date range
    """

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if not start_date or not end_date:
        return JsonResponse({"error": "start_date and end_date required"}, status=400)

    start = timezone.make_aware(
        datetime.combine(
            datetime.strptime(start_date, "%Y-%m-%d").date(),
            datetime.min.time()
        )
    )
    end = timezone.make_aware(
        datetime.combine(
            datetime.strptime(end_date, "%Y-%m-%d").date(),
            datetime.max.time()
        )
    )

    # ---------------- SALES ITEMS ----------------
    sales_items = SalesItem.objects.filter(
        sales_bill__bill_date__range=(start, end)
    )

    # Purchase cost per line
    purchase_cost_expr = ExpressionWrapper(
        F("quantity") * F("purchase_item__billing_price"),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )

    # Aggregate bill-wise
    bill_data = (
        sales_items
        .values(
            "sales_bill_id",
            "sales_bill__bill_number",
            "sales_bill__bill_date",
            "sales_bill__payment_mode",
        )
        .annotate(
            sales_total=Sum("line_total"),
            purchase_cost=Sum(purchase_cost_expr),
        )
        .order_by("-sales_bill__bill_date")
    )

    # ---------------- TOTAL EXPENSES ----------------
    total_expenses = Expense.objects.filter(
        approved=True,
        expense_date__range=(start, end)
    ).aggregate(total=Sum("amount"))["total"] or 0

    total_sales_all_bills = sum(b["sales_total"] for b in bill_data) or 1

    results = []
    for b in bill_data:
        # Proportional expense allocation
        expense_share = (
            (b["sales_total"] / total_sales_all_bills) * total_expenses
        )

        profit = b["sales_total"] - b["purchase_cost"] - expense_share

        results.append({
            "bill_no": b["sales_bill__bill_number"],
            "date": b["sales_bill__bill_date"].strftime("%d-%m-%Y"),
            "payment_mode": b["sales_bill__payment_mode"],
            "sales": float(b["sales_total"]),
            "purchase_cost": float(b["purchase_cost"]),
            "expenses": round(float(expense_share), 2),
            "profit": round(float(profit), 2),
        })

    return JsonResponse({
        "start": start_date,
        "end": end_date,
        "total_expenses": float(total_expenses),
        "bills": results,
    })

@login_required
def bill_profit_dashboard_view(request):
    """
    Renders the Bill-wise Profit Dashboard HTML page
    """
    return render(request, "inventory/bill_profit_dashboard.html")
